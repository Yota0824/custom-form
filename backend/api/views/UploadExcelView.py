from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, serializers
from rest_framework.serializers import Serializer
from openpyxl import load_workbook
from api.models.WebFlyer import WebFlyer

class UploadExcelSerializer(Serializer):
    file = serializers.FileField()
    school_id = serializers.CharField(max_length=4,required=True)
    school_name = serializers.CharField(max_length=20,required=True)
    name = serializers.CharField(max_length=20,required=True)
    apply_date = serializers.CharField(max_length=8,required=True)


class UploadExcelView(APIView):
    def post(self,request):
        data = request.data
        print(  )
        serializer = UploadExcelSerializer(data=data)
        
        if serializer.is_valid():
            file = serializer.validated_data["file"]
            school_id = serializer.validated_data["school_id"] 
            school_name = serializer.validated_data["school_name"] 
            name = serializer.validated_data["name"] 
            apply_date = serializer.validated_data["apply_date"] 
            print(type(apply_date))
            print(apply_date)
            wb = load_workbook(file,data_only=True)
            # sheet = wb.active
            # print(sheet)
            testpath = "エクセル保存処理後何か入る"

            # EXCELの中身取り出して書き込み
            sheet = wb["配布物・WEB"]
            total_out = sheet.cell(8,1).value
            print(total_out)

            # WebFlyer.objects.update_or_create(
            #     # 同一かどうか特定する条件
            #     school_id = school_id,
            #     school_name = school_name,
            #     apply_date = apply_date,
            #     defaults = {
            #         # 更新するもの
            #         "apply_name":name,
            #         "excel_file_path":testpath
            #         }
            # )

            #TODO シート操作
            return Response({"message":"uploaded"},status=status.HTTP_200_OK)
        else:
            return Response(status=400)
        